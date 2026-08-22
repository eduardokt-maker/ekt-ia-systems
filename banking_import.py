from __future__ import annotations

import base64
import csv
import hashlib
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from pypdf import PdfReader


MAX_FILE_BYTES = 10 * 1024 * 1024
SUPPORTED_EXTENSIONS = {".pdf", ".csv", ".txt", ".ofx", ".xlsx"}


def _plain(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(char for char in text if not unicodedata.combining(char)).lower()


def decode_upload(payload: dict[str, Any]) -> tuple[str, bytes]:
    filename = Path(str(payload.get("filename", "arquivo"))).name[:180]
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError("Formato não suportado. Use PDF, CSV, TXT, OFX ou XLSX.")
    try:
        content = base64.b64decode(str(payload.get("content_base64", "")), validate=True)
    except Exception as exc:
        raise ValueError("O arquivo enviado é inválido.") from exc
    if not content or len(content) > MAX_FILE_BYTES:
        raise ValueError("O arquivo deve possuir conteúdo e ter no máximo 10 MB.")
    return filename, content


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def extract_text(filename: str, content: bytes) -> str:
    extension = Path(filename).suffix.lower()
    if extension == ".pdf":
        try:
            reader = PdfReader(io.BytesIO(content))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except Exception as exc:
            raise ValueError("Não foi possível ler este PDF.") from exc
    if extension == ".xlsx":
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    rows.append("\t".join("" if value is None else str(value) for value in row))
            return "\n".join(rows)
        except Exception as exc:
            raise ValueError("Não foi possível interpretar a planilha XLSX.") from exc
    return _decode_text(content)


def _date_text(raw: str) -> str | None:
    raw = raw.strip()
    for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, pattern).date().isoformat()
        except ValueError:
            continue
    match = re.fullmatch(r"(\d{1,2})[/\-](\d{1,2})", raw)
    if match:
        today = date.today()
        try:
            return date(today.year, int(match.group(2)), int(match.group(1))).isoformat()
        except ValueError:
            return None
    return None


def _amount(raw: str) -> Decimal | None:
    cleaned = re.sub(r"[^\d,.-]", "", raw).strip()
    if not cleaned:
        return None
    negative = cleaned.startswith("-") or cleaned.endswith("-")
    cleaned = cleaned.strip("-")
    if "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        value = Decimal(cleaned)
    except InvalidOperation:
        return None
    return -value if negative else value


def _money(value: Decimal) -> str:
    return f"{abs(value):.2f}".replace(".", ",")


def _type(description: str, value: Decimal, document_kind: str) -> str:
    plain = _plain(description)
    incoming = ("recebido", "credito", "salario", "deposito", "resgate", "estorno", "entrada")
    outgoing = ("enviado", "debito", "pagamento", "compra", "tarifa", "saque", "saida")
    if value < 0 or any(word in plain for word in outgoing):
        return "EXPENSE"
    if value > 0 or any(word in plain for word in incoming):
        return "INCOME"
    return "EXPENSE" if document_kind == "receipt" else "INCOME"


def _category_hint(description: str, transaction_type: str) -> str:
    plain = _plain(description)
    rules = (
        (("salario",), "Salário"), (("aluguel",), "Aluguel"),
        (("pix recebido",), "PIX recebido"), (("tarifa",), "Tarifas bancárias"),
        (("juros",), "Juros"), (("posto", "combustivel", "shell"), "Combustível"),
        (("supermercado", "carrefour"), "Supermercado"),
        (("ifood", "delivery"), "Delivery"), (("uber",), "Transporte"),
        (("farmacia", "drogasil"), "Farmácia"),
        (("netflix", "spotify", "streaming"), "Assinaturas"),
    )
    for words, category in rules:
        if any(word in plain for word in words):
            return category
    return "Outros"


def _detected_bank(text: str) -> str:
    banks = ("Itaú", "Nubank", "Banco do Brasil", "Bradesco", "Santander", "Caixa", "Inter", "Sicredi")
    plain = _plain(text)
    return next((bank for bank in banks if _plain(bank) in plain), "")


def _candidate(tx_date: str, description: str, value: Decimal, document_kind: str, source_line: str) -> dict[str, Any]:
    transaction_type = _type(description, value, document_kind)
    fingerprint = hashlib.sha256(f"{tx_date}|{transaction_type}|{abs(value):.2f}|{_plain(description)}".encode()).hexdigest()[:24]
    return {
        "selected": True,
        "transaction_date": tx_date,
        "reference_month": tx_date[:7],
        "transaction_type": transaction_type,
        "description": description.strip()[:140] or "Movimentação importada",
        "counterparty": description.strip()[:140],
        "amount": _money(value),
        "payment_method": "Importado automaticamente",
        "category_hint": _category_hint(description, transaction_type),
        "external_id": fingerprint,
        "source_line": source_line.strip()[:500],
    }


def _parse_ofx(text: str, document_kind: str) -> list[dict[str, Any]]:
    items = []
    for block in re.findall(r"<STMTTRN>(.*?)(?=<STMTTRN>|</BANKTRANLIST>|$)", text, flags=re.I | re.S):
        def tag(name: str) -> str:
            match = re.search(fr"<{name}>([^<\r\n]+)", block, flags=re.I)
            return match.group(1).strip() if match else ""
        raw_date, raw_amount = tag("DTPOSTED")[:8], tag("TRNAMT")
        try:
            tx_date = datetime.strptime(raw_date, "%Y%m%d").date().isoformat()
            value = Decimal(raw_amount.replace(",", "."))
        except (ValueError, InvalidOperation):
            continue
        description = " ".join(part for part in (tag("NAME"), tag("MEMO")) if part) or "Movimentação OFX"
        item = _candidate(tx_date, description, value, document_kind, block)
        item["external_id"] = tag("FITID") or item["external_id"]
        items.append(item)
    return items


def _parse_csv(text: str, document_kind: str) -> list[dict[str, Any]]:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    if not rows:
        return []
    headers = [_plain(cell).strip() for cell in rows[0]]
    date_index = next((i for i, value in enumerate(headers) if "data" in value or "date" in value), 0)
    description_index = next((i for i, value in enumerate(headers) if any(word in value for word in ("descricao", "historico", "memo", "estabelecimento"))), 1 if len(headers) > 1 else 0)
    amount_index = next((i for i, value in enumerate(headers) if any(word in value for word in ("valor", "amount", "debito", "credito"))), len(headers) - 1)
    items = []
    for row in rows[1:]:
        if max(date_index, description_index, amount_index) >= len(row):
            continue
        tx_date, value = _date_text(row[date_index]), _amount(row[amount_index])
        if tx_date and value is not None and value != 0:
            items.append(_candidate(tx_date, row[description_index], value, document_kind, " | ".join(row)))
    return items


def _parse_lines(text: str, document_kind: str) -> list[dict[str, Any]]:
    items = []
    date_pattern = re.compile(r"\b(\d{1,2}[/\-]\d{1,2}(?:[/\-]\d{2,4})?|\d{4}-\d{2}-\d{2})\b")
    amount_pattern = re.compile(r"(?:R\$\s*)?[-+]?\s*\d{1,3}(?:\.\d{3})*,\d{2}-?|(?:R\$\s*)?[-+]?\s*\d+[.,]\d{2}-?")
    for line in (part.strip() for part in text.splitlines()):
        date_match = date_pattern.search(line)
        amounts = list(amount_pattern.finditer(line))
        if not date_match or not amounts:
            continue
        tx_date = _date_text(date_match.group(1))
        value = _amount(amounts[-1].group(0))
        if not tx_date or value is None or value == 0:
            continue
        description = (line[:amounts[-1].start()] + line[amounts[-1].end():]).replace(date_match.group(0), " ")
        description = re.sub(r"\s+", " ", description).strip(" -|:")
        items.append(_candidate(tx_date, description, value, document_kind, line))
    return items


def preview(payload: dict[str, Any]) -> dict[str, Any]:
    filename, content = decode_upload(payload)
    kind = str(payload.get("document_kind", "statement"))
    if kind not in {"statement", "receipt"}:
        raise ValueError("Tipo de documento inválido.")
    text = extract_text(filename, content).strip()
    if not text:
        raise ValueError("O arquivo não possui texto pesquisável. Envie um PDF com texto ou utilize OCR antes da importação.")
    extension = Path(filename).suffix.lower()
    if extension == ".ofx":
        items = _parse_ofx(text, kind)
    elif extension in {".csv", ".xlsx"}:
        items = _parse_csv(text, kind)
    else:
        items = _parse_lines(text, kind)
    if kind == "receipt" and not items:
        date_match = re.search(r"\b(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}|\d{4}-\d{2}-\d{2})\b", text)
        amount_match = re.search(r"valor\s*:?\s*(?:R\$\s*)?([-+]?\s*\d{1,3}(?:\.\d{3})*,\d{2}|[-+]?\s*\d+[.,]\d{2})", text, flags=re.I)
        if amount_match is None:
            amount_match = re.search(r"R\$\s*([-+]?\s*\d{1,3}(?:\.\d{3})*,\d{2}|[-+]?\s*\d+[.,]\d{2})", text, flags=re.I)
        if date_match and amount_match:
            tx_date, value = _date_text(date_match.group(1)), _amount(amount_match.group(1))
            description_lines = [line.strip() for line in text.splitlines() if any(word in _plain(line) for word in ("favorecido", "recebedor", "pagador", "pix", "ted", "transferencia"))]
            description = " • ".join(description_lines[:2]) or "Comprovante bancário"
            if tx_date and value is not None:
                items = [_candidate(tx_date, description, value, kind, text)]
    if kind == "receipt" and len(items) > 1:
        items = items[:1]
    if not items:
        raise ValueError("Nenhuma movimentação pôde ser identificada automaticamente. Confira se o arquivo contém data e valor legíveis.")
    return {
        "ok": True,
        "filename": filename,
        "document_kind": kind,
        "detected_bank": _detected_bank(text),
        "items": items,
        "notice": "Revise os dados. A leitura automática é uma sugestão e nada foi salvo ainda.",
    }
